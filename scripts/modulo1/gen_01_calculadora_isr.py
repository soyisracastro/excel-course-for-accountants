"""
Generador: 01_Calculadora_ISR_V2026.xlsx
Módulo 1 — Lógica Contable y Funciones de Control

Hojas:
  - Tarifa_Anual_2026: Tabla ISR anual completa
  - Calculadora_ISR: Cálculo con BUSCARV paso a paso
  - Actualizacion_CFF: Factor de actualización con INPC + TRUNCAR
  - Ejercicios: 5 escenarios de cálculo
  - Instrucciones
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl.utils import get_column_letter
from scripts.config.constants import PACK, Color
from scripts.config.isr_2026 import TARIFA_ANUAL, TARIFA_MENSUAL, INPC_RECIENTE, INPC_ANTERIOR
from scripts.config.styles import (
    FILL_HEADER, FILL_LIGHT, FILL_VERDE, FILL_AMARILLO,
    FONT_HEADER, FONT_TITULO_XL, FONT_SUBTITULO, FONT_NORMAL, FONT_SMALL,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_RIGHT,
    FMT_MONEY, FMT_PCT, FMT_MONEY_4DEC,
    apply_header_style, apply_data_style, style_title_cell, auto_width
)
from scripts.generators.xlsx_gen import ExcelGenerator

OUTPUT_DIR = PACK / "Modulo_1_Funciones"


def build():
    gen = ExcelGenerator("01_Calculadora_ISR_V2026.xlsx", OUTPUT_DIR)

    # ── Hoja 1: Tarifa Anual 2026 ────────────────────────────────
    ws1 = gen.add_sheet("Tarifa_Anual_2026")
    style_title_cell(ws1, 1, 1, "Tarifa ISR Anual 2026 — Art. 152 LISR (Anexo 8 RMF)", 6)

    headers = ["Límite Inferior", "Límite Superior", "Cuota Fija", "% Sobre Excedente"]
    data = []
    for r in TARIFA_ANUAL:
        sup = "En adelante" if r["lim_sup"] > 999_999_999 else r["lim_sup"]
        data.append([r["lim_inf"], sup, r["cuota"], r["pct"] / 100])

    gen.write_table(ws1, headers, data, start_row=3, table_name="Tarifa_Anual_2026",
                    money_cols=[1, 2, 3], pct_cols=[4])

    # Tarifa mensual
    ws1m = gen.add_sheet("Tarifa_Mensual_2026")
    style_title_cell(ws1m, 1, 1, "Tarifa ISR Mensual 2026 — Art. 96 LISR (Anexo 8 RMF)", 6)

    data_m = []
    for r in TARIFA_MENSUAL:
        sup = "En adelante" if r["lim_sup"] > 999_999_999 else r["lim_sup"]
        data_m.append([r["lim_inf"], sup, r["cuota"], r["pct"] / 100])

    gen.write_table(ws1m, headers, data_m, start_row=3, table_name="Tarifa_Mensual_2026",
                    money_cols=[1, 2, 3], pct_cols=[4])

    # ── Hoja 2: Calculadora ISR ──────────────────────────────────
    ws2 = gen.add_sheet("Calculadora_ISR")
    style_title_cell(ws2, 1, 1, "Calculadora de ISR Anual 2026", 5)
    ws2.cell(row=2, column=1, value="Ingresa tu base gravable en la celda B4 y observa cómo las fórmulas calculan el ISR automáticamente.").font = FONT_SMALL

    labels = [
        (4, "Base Gravable (ingreso acumulable - deducciones)"),
        (5, "Límite Inferior (BUSCARV)"),
        (6, "Excedente sobre Límite Inferior"),
        (7, "% Sobre Excedente (BUSCARV)"),
        (8, "ISR Marginal"),
        (9, "Cuota Fija (BUSCARV)"),
        (10, "ISR del Ejercicio"),
    ]

    for row, label in labels:
        ws2.cell(row=row, column=1, value=label).font = FONT_NORMAL
        ws2.cell(row=row, column=1).border = THIN_BORDER
        ws2.cell(row=row, column=2).border = THIN_BORDER
        ws2.cell(row=row, column=2).number_format = FMT_MONEY
        ws2.cell(row=row, column=2).alignment = ALIGN_RIGHT

    # Input cell
    ws2.cell(row=4, column=2, value=450000).font = FONT_SUBTITULO
    ws2.cell(row=4, column=2).fill = FILL_AMARILLO

    # Formulas (English — Excel translates to Spanish)
    ws2["B5"] = '=VLOOKUP(B4,Tarifa_Anual_2026,1,TRUE)'
    ws2["B6"] = '=B4-B5'
    ws2["B7"] = '=VLOOKUP(B4,Tarifa_Anual_2026,4,TRUE)'
    ws2["B7"].number_format = FMT_PCT
    ws2["B8"] = '=B6*B7'
    ws2["B9"] = '=VLOOKUP(B4,Tarifa_Anual_2026,3,TRUE)'
    ws2["B10"] = '=B8+B9'
    ws2["B10"].fill = FILL_VERDE
    ws2["B10"].font = FONT_SUBTITULO

    # Explanations
    explanations = {
        5: "BUSCARV busca el límite inferior que corresponde a tu ingreso",
        6: "Tu ingreso menos el límite inferior de tu rango",
        7: "El porcentaje de impuesto sobre el excedente",
        8: "Excedente × Porcentaje = ISR marginal",
        9: "Cantidad fija que se suma según tu rango",
        10: "ISR Marginal + Cuota Fija = ISR Total del ejercicio",
    }
    for row, text in explanations.items():
        ws2.cell(row=row, column=3, value=text).font = FONT_SMALL

    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 55

    # ── Hoja 3: Factor de Actualización ──────────────────────────
    ws3 = gen.add_sheet("Actualizacion_CFF")
    style_title_cell(ws3, 1, 1, "Factor de Actualización — Art. 17-A CFF", 5)
    ws3.cell(row=2, column=1, value="El factor se trunca a 4 decimales (diezmilésimo) conforme al CFF.").font = FONT_SMALL

    labels3 = [
        (4, "INPC Reciente (Dic 2025)"),
        (5, "INPC Anterior (Dic 2024)"),
        (6, "Factor de Actualización (sin truncar)"),
        (7, "Factor de Actualización (TRUNCAR a 4 dec)"),
        (8, ""),
        (9, "Monto Original"),
        (10, "Monto Actualizado"),
    ]

    for row, label in labels3:
        if label:
            ws3.cell(row=row, column=1, value=label).font = FONT_NORMAL
            ws3.cell(row=row, column=1).border = THIN_BORDER
            ws3.cell(row=row, column=2).border = THIN_BORDER

    ws3["B4"] = INPC_RECIENTE
    ws3["B5"] = INPC_ANTERIOR
    ws3["B6"] = "=B4/B5"
    ws3["B6"].number_format = "0.000000"
    ws3["B7"] = "=TRUNC(B4/B5,4)"
    ws3["B7"].number_format = FMT_MONEY_4DEC
    ws3["B7"].fill = FILL_VERDE

    ws3["B9"] = 100000
    ws3["B9"].fill = FILL_AMARILLO
    ws3["B9"].number_format = FMT_MONEY
    ws3["B10"] = "=B9*B7"
    ws3["B10"].number_format = FMT_MONEY
    ws3["B10"].fill = FILL_VERDE

    ws3.cell(row=6, column=3, value="División simple INPC reciente / INPC anterior").font = FONT_SMALL
    ws3.cell(row=7, column=3, value="TRUNCAR(valor, 4) — trunca sin redondear, como exige el CFF").font = FONT_SMALL
    ws3.cell(row=10, column=3, value="Monto × Factor = Monto actualizado por inflación").font = FONT_SMALL

    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 55

    # ── Hoja 4: Ejercicios ───────────────────────────────────────
    ws4 = gen.add_sheet("Ejercicios")
    style_title_cell(ws4, 1, 1, "Ejercicios de Cálculo ISR 2026", 6)
    ws4.cell(row=2, column=1, value="Calcula el ISR para cada escenario. Usa BUSCARV con la tarifa anual.").font = FONT_SMALL

    headers_ej = ["#", "Escenario", "Ingreso Anual", "Deducciones", "Base Gravable", "ISR (tu respuesta)", "ISR (verificación)"]
    ejercicios = [
        [1, "Empleado con sueldo fijo", 280000, 45000, None, None, None],
        [2, "Freelancer con ingresos variables", 520000, 120000, None, None, None],
        [3, "Socio de empresa (dividendos)", 1500000, 350000, None, None, None],
        [4, "Trabajador zona fronteriza", 180000, 30000, None, None, None],
        [5, "Director general (ingreso alto)", 3200000, 600000, None, None, None],
    ]

    gen.write_table(ws4, headers_ej, ejercicios, start_row=4, money_cols=[3, 4, 5, 6, 7])

    # Add formulas for Base Gravable and ISR verification
    for i in range(len(ejercicios)):
        row = 5 + i
        ws4.cell(row=row, column=5).value = f"=C{row}-D{row}"
        ws4.cell(row=row, column=7).value = (
            f'=VLOOKUP(E{row},Tarifa_Anual_2026,3,TRUE)'
            f'+(E{row}-VLOOKUP(E{row},Tarifa_Anual_2026,1,TRUE))'
            f'*VLOOKUP(E{row},Tarifa_Anual_2026,4,TRUE)'
        )

    # ── Hoja 5: Instrucciones ────────────────────────────────────
    gen.add_instructions_sheet([
        "Este archivo contiene la tarifa ISR 2026 oficial (Anexo 8 RMF, DOF 28-dic-2025).",
        "En la hoja 'Calculadora_ISR', cambia el valor amarillo (B4) para calcular automáticamente el ISR.",
        "Las fórmulas usan BUSCARV (VLOOKUP en inglés) para buscar el rango correcto en la tarifa.",
        "En 'Actualizacion_CFF', observa cómo TRUNCAR elimina decimales sin redondear (Art. 17-A CFF).",
        "Completa los 5 ejercicios en la hoja 'Ejercicios'. La columna G tiene la verificación.",
        "IMPORTANTE: Excel muestra las funciones en español (BUSCARV, SI, TRUNCAR) aunque se programaron en inglés.",
        "Consejo: Usa F2 para entrar a una celda y ver la fórmula en la barra de fórmulas.",
    ])

    gen.save()


if __name__ == "__main__":
    build()
