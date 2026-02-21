"""
Generador: 02_Control_Vencimientos_EFirma.xlsx
Módulo 1 — Control de vencimiento de e.firma (firma electrónica SAT)

10 empresas sample con:
  - =fecha-HOY() para días restantes
  - SI anidado con semáforo: VIGENTE / POR CADUCAR / CADUCADA
  - Formato condicional verde/amarillo/rojo
"""
import sys
from pathlib import Path
from datetime import date, timedelta
import random

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl.formatting.rule import CellIsRule
from scripts.config.constants import PACK, Color
from scripts.config.styles import (
    FILL_VERDE, FILL_AMARILLO, FILL_ROJO, FILL_LIGHT,
    FONT_HEADER, FONT_NORMAL, FONT_SMALL, FONT_TITULO_XL,
    THIN_BORDER, ALIGN_CENTER, FMT_DATE,
    apply_header_style, apply_data_style, style_title_cell, auto_width
)
from scripts.generators.xlsx_gen import ExcelGenerator

OUTPUT_DIR = PACK / "Modulo_1_Funciones"

EMPRESAS = [
    ("Comercializadora del Norte SA de CV", "CNO850315XX1"),
    ("Grupo Contable Torres SC", "GCT920801KL3"),
    ("Distribuidora Azteca SA de CV", "DAZ100512M45"),
    ("Servicios Administrativos Mora SC", "SAM880620PQ2"),
    ("Industrias Metalúrgicas del Bajío SA", "IMB760910RS7"),
    ("Consultores Fiscales Hernández SC", "CFH051103TU8"),
    ("Transportes y Logística Reyes SA de CV", "TLR110215VW9"),
    ("Farmacia Santa Cruz SA de CV", "FSC990430XY0"),
    ("Constructora e Inmobiliaria Vega SA", "CIV070817ZA1"),
    ("Despacho Contable Martínez y Asociados", "DCM130225BC4"),
]


def build():
    gen = ExcelGenerator("02_Control_Vencimientos_EFirma.xlsx", OUTPUT_DIR)
    ws = gen.add_sheet("Control_EFirma")

    style_title_cell(ws, 1, 1, "Control de Vencimiento de e.firma (Firma Electrónica)", 7)
    ws.cell(row=2, column=1, value="Monitorea la vigencia de la e.firma de tus clientes con funciones SI, HOY y aritmética de fechas.").font = FONT_SMALL

    headers = [
        "Empresa", "RFC", "Fecha Emisión", "Fecha Vencimiento",
        "Días Restantes", "Estatus", "Observaciones"
    ]

    random.seed(42)
    hoy = date(2026, 2, 21)
    data = []
    for nombre, rfc in EMPRESAS:
        # Generate varied dates: some expired, some about to expire, some ok
        emision = hoy - timedelta(days=random.randint(200, 1400))
        vigencia_dias = 365 * random.choice([2, 4])
        vencimiento = emision + timedelta(days=vigencia_dias)
        data.append([nombre, rfc, emision, vencimiento, None, None, ""])

    last_row = gen.write_table(ws, headers, data, start_row=4,
                               date_cols=[3, 4])

    # Add formulas for Días Restantes and Estatus
    for i in range(len(data)):
        row = 5 + i
        # Días restantes = Fecha Vencimiento - HOY()
        ws.cell(row=row, column=5).value = f"=D{row}-TODAY()"
        ws.cell(row=row, column=5).number_format = "0"
        ws.cell(row=row, column=5).alignment = ALIGN_CENTER

        # SI anidado: > 90 → VIGENTE, > 0 → POR CADUCAR, else CADUCADA
        ws.cell(row=row, column=6).value = (
            f'=IF(E{row}>90,"VIGENTE",IF(E{row}>0,"POR CADUCAR","CADUCADA"))'
        )
        ws.cell(row=row, column=6).alignment = ALIGN_CENTER

    # Conditional formatting on Estatus column (F)
    from openpyxl.styles import PatternFill, Font
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(color="006100")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    yellow_font = Font(color="9C6500")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006")

    range_str = f"F5:F{4 + len(data)}"
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="equal", formula=['"VIGENTE"'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="equal", formula=['"POR CADUCAR"'], fill=yellow_fill, font=yellow_font))
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="equal", formula=['"CADUCADA"'], fill=red_fill, font=red_font))

    # Also conditional on Días Restantes column (E)
    range_dias = f"E5:E{4 + len(data)}"
    ws.conditional_formatting.add(range_dias,
        CellIsRule(operator="greaterThan", formula=["90"], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(range_dias,
        CellIsRule(operator="between", formula=["1", "90"], fill=yellow_fill, font=yellow_font))
    ws.conditional_formatting.add(range_dias,
        CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=red_fill, font=red_font))

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 25

    # Instructions sheet
    gen.add_instructions_sheet([
        "Este archivo monitorea la vigencia de la e.firma (firma electrónica) del SAT.",
        "La columna 'Días Restantes' usa la fórmula: =FechaVencimiento - HOY()",
        "La columna 'Estatus' usa SI anidado: SI(días>90, 'VIGENTE', SI(días>0, 'POR CADUCAR', 'CADUCADA'))",
        "Los colores del semáforo se aplican con Formato Condicional automático.",
        "Verde = VIGENTE (más de 90 días), Amarillo = POR CADUCAR (1-90 días), Rojo = CADUCADA.",
        "Agrega tus propios clientes en las filas siguientes y copia las fórmulas.",
        "La función HOY() se actualiza cada vez que abres el archivo — siempre ves datos al día.",
    ])

    gen.save()


if __name__ == "__main__":
    build()
