"""
Generador: 03_Extraccion_RFC_Master.xlsx
Módulo 1 — Extracción de componentes del RFC con EXTRAE (MID)

15 RFCs ficticios pero estructuralmente válidos:
  - PF = 13 caracteres, PM = 12 caracteres
  - EXTRAE por posición para extraer letras, año, mes, día
  - FECHA para reconstruir fecha de nacimiento/constitución
  - Comparar con HOY() para calcular edad
"""
import sys
from pathlib import Path
import random

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from scripts.config.constants import PACK
from scripts.config.styles import (
    FILL_AMARILLO, FILL_VERDE, FILL_LIGHT,
    FONT_NORMAL, FONT_SMALL, FONT_SUBTITULO,
    THIN_BORDER, ALIGN_CENTER, FMT_DATE,
    style_title_cell, auto_width
)
from scripts.generators.xlsx_gen import ExcelGenerator

OUTPUT_DIR = PACK / "Modulo_1_Funciones"

# RFCs ficticios pero estructuralmente válidos
RFCS_PF = [  # Persona Física: 4 letras + 6 dígitos fecha + 3 homoclave = 13
    ("Juan Pérez López", "PELJ850312AB1"),
    ("María García Hernández", "GAHM900725CD2"),
    ("Carlos Rodríguez Martínez", "ROMC780118EF3"),
    ("Ana López Sánchez", "LOSA951230GH4"),
    ("Pedro Díaz Torres", "DITP880615IJ5"),
    ("Laura Fernández Ruiz", "FERL920403KL6"),
    ("Miguel Castro Flores", "CAFM010520MN7"),
    ("Sofía Ramírez Vega", "RAVS990108OP8"),
    ("Roberto Morales Cruz", "MOCR831127QR9"),
    ("Elena Ortiz Jiménez", "OIJE870214ST0"),
]

RFCS_PM = [  # Persona Moral: 3 letras + 6 dígitos fecha + 3 homoclave = 12
    ("Comercializadora del Pacífico SA de CV", "CPA150610UV1"),
    ("Grupo Industrial Norteño SA", "GIN080923WX2"),
    ("Servicios Digitales Mx SC", "SDM200315YZ3"),
    ("Transportes Nacionales SA de CV", "TNA170801AB4"),
    ("Constructora Vega y Asociados SC", "CVA190412CD5"),
]


def build():
    gen = ExcelGenerator("03_Extraccion_RFC_Master.xlsx", OUTPUT_DIR)

    # ── Hoja 1: Extracción RFC ───────────────────────────────────
    ws = gen.add_sheet("Extraccion_RFC")
    style_title_cell(ws, 1, 1, "Extracción de Componentes del RFC con EXTRAE", 10)
    ws.cell(row=2, column=1,
            value="Usa EXTRAE (MID en inglés) para descomponer el RFC en sus partes. PF=13 chars, PM=12 chars.").font = FONT_SMALL

    headers = [
        "Nombre/Razón Social", "RFC", "Tipo", "Letras",
        "Año (2 díg)", "Mes", "Día",
        "Año Completo", "Fecha Reconstruida", "Edad/Antigüedad",
        "Homoclave"
    ]

    all_rfcs = [(n, r, "PF") for n, r in RFCS_PF] + [(n, r, "PM") for n, r in RFCS_PM]
    data = [[n, r, t] + [None] * 8 for n, r, t in all_rfcs]

    gen.write_table(ws, headers, data, start_row=4)

    # Add EXTRAE formulas
    for i in range(len(all_rfcs)):
        row = 5 + i
        tipo = all_rfcs[i][2]

        if tipo == "PF":
            # PF: 4 letras (pos 1-4), fecha (pos 5-10), homoclave (pos 11-13)
            ws.cell(row=row, column=4).value = f'=MID(B{row},1,4)'       # Letras
            ws.cell(row=row, column=5).value = f'=MID(B{row},5,2)'       # Año
            ws.cell(row=row, column=6).value = f'=MID(B{row},7,2)'       # Mes
            ws.cell(row=row, column=7).value = f'=MID(B{row},9,2)'       # Día
            ws.cell(row=row, column=11).value = f'=MID(B{row},11,3)'     # Homoclave
        else:
            # PM: 3 letras (pos 1-3), fecha (pos 4-9), homoclave (pos 10-12)
            ws.cell(row=row, column=4).value = f'=MID(B{row},1,3)'
            ws.cell(row=row, column=5).value = f'=MID(B{row},4,2)'
            ws.cell(row=row, column=6).value = f'=MID(B{row},6,2)'
            ws.cell(row=row, column=7).value = f'=MID(B{row},8,2)'
            ws.cell(row=row, column=11).value = f'=MID(B{row},10,3)'

        # Año completo: si > 30, es 1900s; si <= 30, es 2000s
        ws.cell(row=row, column=8).value = (
            f'=IF(VALUE(E{row})>30,1900+VALUE(E{row}),2000+VALUE(E{row}))'
        )

        # Fecha reconstruida: FECHA(año, mes, día)
        ws.cell(row=row, column=9).value = (
            f'=DATE(H{row},VALUE(F{row}),VALUE(G{row}))'
        )
        ws.cell(row=row, column=9).number_format = FMT_DATE

        # Edad/Antigüedad en años
        ws.cell(row=row, column=10).value = (
            f'=INT((TODAY()-I{row})/365.25)'
        )
        ws.cell(row=row, column=10).alignment = ALIGN_CENTER

    # Adjust widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 6
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 12

    # ── Hoja 2: Diagrama de posiciones ───────────────────────────
    ws2 = gen.add_sheet("Posiciones_RFC")
    style_title_cell(ws2, 1, 1, "Diagrama de Posiciones del RFC", 14)

    ws2.cell(row=3, column=1, value="Persona Física (13 caracteres):").font = FONT_SUBTITULO
    pf_headers = ["Pos 1", "Pos 2", "Pos 3", "Pos 4", "Pos 5", "Pos 6",
                   "Pos 7", "Pos 8", "Pos 9", "Pos 10", "Pos 11", "Pos 12", "Pos 13"]
    pf_desc = ["Letra", "Letra", "Letra", "Letra", "Año", "Año",
               "Mes", "Mes", "Día", "Día", "Homo", "Homo", "Homo"]
    pf_example = ["P", "E", "L", "J", "8", "5", "0", "3", "1", "2", "A", "B", "1"]

    for i, (h, d, e) in enumerate(zip(pf_headers, pf_desc, pf_example)):
        col = i + 1
        ws2.cell(row=4, column=col, value=h).font = FONT_SMALL
        ws2.cell(row=4, column=col).alignment = ALIGN_CENTER
        ws2.cell(row=4, column=col).border = THIN_BORDER
        ws2.cell(row=5, column=col, value=d).font = FONT_SMALL
        ws2.cell(row=5, column=col).alignment = ALIGN_CENTER
        ws2.cell(row=5, column=col).border = THIN_BORDER
        ws2.cell(row=6, column=col, value=e).font = FONT_NORMAL
        ws2.cell(row=6, column=col).alignment = ALIGN_CENTER
        ws2.cell(row=6, column=col).border = THIN_BORDER

        # Color code groups
        from openpyxl.styles import PatternFill
        if i < 4:
            ws2.cell(row=6, column=col).fill = PatternFill("solid", fgColor="DBEAFE")
        elif i < 10:
            ws2.cell(row=6, column=col).fill = PatternFill("solid", fgColor="D1FAE5")
        else:
            ws2.cell(row=6, column=col).fill = PatternFill("solid", fgColor="FEF3C7")

    ws2.cell(row=8, column=1, value="Persona Moral (12 caracteres):").font = FONT_SUBTITULO
    pm_headers = pf_headers[:12]
    pm_desc = ["Letra", "Letra", "Letra", "Año", "Año",
               "Mes", "Mes", "Día", "Día", "Homo", "Homo", "Homo"]
    pm_example = ["C", "P", "A", "1", "5", "0", "6", "1", "0", "U", "V", "1"]

    for i, (h, d, e) in enumerate(zip(pm_headers, pm_desc, pm_example)):
        col = i + 1
        ws2.cell(row=9, column=col, value=h).font = FONT_SMALL
        ws2.cell(row=9, column=col).alignment = ALIGN_CENTER
        ws2.cell(row=9, column=col).border = THIN_BORDER
        ws2.cell(row=10, column=col, value=d).font = FONT_SMALL
        ws2.cell(row=10, column=col).alignment = ALIGN_CENTER
        ws2.cell(row=10, column=col).border = THIN_BORDER
        ws2.cell(row=11, column=col, value=e).font = FONT_NORMAL
        ws2.cell(row=11, column=col).alignment = ALIGN_CENTER
        ws2.cell(row=11, column=col).border = THIN_BORDER

        if i < 3:
            ws2.cell(row=11, column=col).fill = PatternFill("solid", fgColor="DBEAFE")
        elif i < 9:
            ws2.cell(row=11, column=col).fill = PatternFill("solid", fgColor="D1FAE5")
        else:
            ws2.cell(row=11, column=col).fill = PatternFill("solid", fgColor="FEF3C7")

    ws2.cell(row=13, column=1, value="Azul = Letras identificadoras").font = FONT_SMALL
    ws2.cell(row=14, column=1, value="Verde = Fecha (AAMMDD)").font = FONT_SMALL
    ws2.cell(row=15, column=1, value="Amarillo = Homoclave (asignada por SAT)").font = FONT_SMALL

    # Instructions
    gen.add_instructions_sheet([
        "Este archivo practica la función EXTRAE (MID) para descomponer el RFC.",
        "PF (Persona Física) tiene 13 caracteres: 4 letras + 6 dígitos fecha + 3 homoclave.",
        "PM (Persona Moral) tiene 12 caracteres: 3 letras + 6 dígitos fecha + 3 homoclave.",
        "La hoja 'Posiciones_RFC' muestra un diagrama visual de las posiciones.",
        "EXTRAE(texto, posición_inicial, número_caracteres) extrae caracteres de un texto.",
        "La fecha se reconstruye con FECHA(año, mes, día) y se compara con HOY() para calcular edad.",
        "Los RFCs en este archivo son ficticios pero siguen la estructura real del SAT.",
    ])

    gen.save()


if __name__ == "__main__":
    build()
