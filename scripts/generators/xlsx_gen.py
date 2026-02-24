"""
Clase base para generar archivos Excel (.xlsx) con openpyxl.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from scripts.config.styles import (
    apply_header_style, apply_data_style, style_title_cell,
    style_instruction_cell, auto_width, FILL_LIGHT,
    FONT_NORMAL, THIN_BORDER, ALIGN_CENTER,
    FMT_MONEY, FMT_PCT, FMT_DATE
)


class ExcelGenerator:
    """Base class for generating Excel workbooks."""

    def __init__(self, filename: str, output_dir: Path):
        self.filename = filename
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.wb = Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)

    @property
    def filepath(self) -> Path:
        return self.output_dir / self.filename

    def add_sheet(self, name: str):
        """Add a new worksheet and return it."""
        return self.wb.create_sheet(title=name)

    def write_table(self, ws, headers, data, start_row=1, start_col=1,
                    table_name=None, money_cols=None, pct_cols=None, date_cols=None):
        """Write headers + data rows with styling, optionally as a named Table."""
        col_end = start_col + len(headers) - 1

        # Headers
        for i, h in enumerate(headers):
            ws.cell(row=start_row, column=start_col + i, value=h)

        # Data
        for r_idx, row_data in enumerate(data):
            row_num = start_row + 1 + r_idx
            for c_idx, val in enumerate(row_data):
                ws.cell(row=row_num, column=start_col + c_idx, value=val)

        if table_name:
            # Named Table: the table style handles header formatting and
            # zebra striping. Only apply number formats to data cells.
            from openpyxl.utils import get_column_letter
            money_cols = money_cols or []
            pct_cols = pct_cols or []
            date_cols = date_cols or []
            for r_idx in range(len(data)):
                row_num = start_row + 1 + r_idx
                for col in range(start_col, col_end + 1):
                    cell = ws.cell(row=row_num, column=col)
                    if col in money_cols:
                        cell.number_format = FMT_MONEY
                    elif col in pct_cols:
                        cell.number_format = FMT_PCT
                    elif col in date_cols:
                        cell.number_format = FMT_DATE

            ref = (f"{get_column_letter(start_col)}{start_row}:"
                   f"{get_column_letter(col_end)}{start_row + len(data)}")
            tbl = Table(displayName=table_name, ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(tbl)
        else:
            # No named table: apply full manual styling
            apply_header_style(ws, start_row, start_col, col_end)
            for r_idx, row_data in enumerate(data):
                row_num = start_row + 1 + r_idx
                apply_data_style(ws, row_num, start_col, col_end,
                                 money_cols=money_cols, pct_cols=pct_cols,
                                 date_cols=date_cols)
            # Zebra striping
            for r_idx in range(0, len(data), 2):
                row_num = start_row + 1 + r_idx
                for col in range(start_col, col_end + 1):
                    ws.cell(row=row_num, column=col).fill = FILL_LIGHT

        auto_width(ws)
        return start_row + len(data)  # last data row

    def add_instructions_sheet(self, instructions, sheet_name="Instrucciones"):
        """Add a sheet with numbered instructions."""
        ws = self.add_sheet(sheet_name)
        style_title_cell(ws, 1, 1, f"📋 {sheet_name}", merge_end_col=8)
        ws.row_dimensions[1].height = 30
        for i, text in enumerate(instructions, start=1):
            row = i + 2
            ws.cell(row=row, column=1, value=i).font = FONT_NORMAL
            ws.cell(row=row, column=1).alignment = ALIGN_CENTER
            # Prefix with space to prevent Excel interpreting as formula
            safe_text = " " + text if text.startswith("=") else text
            cell = ws.cell(row=row, column=2, value=safe_text)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER.__class__(
                horizontal="left", vertical="top", wrap_text=True
            )
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            ws.row_dimensions[row].height = max(30, len(text) // 3)
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 80

    def save(self):
        """Save workbook to disk."""
        self.wb.save(self.filepath)
        print(f"  ✓ {self.filepath.name}")
        return self.filepath
