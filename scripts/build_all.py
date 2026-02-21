"""
build_all.py — Orquestador maestro del curso "Excel para Contadores y Administrativos"
Ejecuta todos los generadores y verifica los outputs.

Uso:
    python3 -m scripts.build_all          # Genera todo
    python3 -m scripts.build_all --verify  # Solo verificar outputs existentes
"""
import sys
import importlib
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.config.constants import PACK, SLIDES_DIR, TELEPROMPTER_DIR

# ── Generadores por fase ─────────────────────────────────────────
GENERATORS = [
    # Fase 2: Módulo 1
    ("scripts.modulo1.gen_01_calculadora_isr", "Módulo 1: Calculadora ISR"),
    ("scripts.modulo1.gen_02_control_vencimientos", "Módulo 1: Control Vencimientos"),
    ("scripts.modulo1.gen_03_extraccion_rfc", "Módulo 1: Extracción RFC"),
    ("scripts.modulo1.gen_slides_m1", "Módulo 1: Slides"),
    ("scripts.modulo1.gen_pdf_ref_m1", "Módulo 1: PDF Referencia"),
    # Fase 3: Módulo 2
    ("scripts.modulo2.gen_04_limpieza_masiva", "Módulo 2: Limpieza Masiva"),
    ("scripts.modulo2.gen_05_nomina_xml", "Módulo 2: Nómina XML"),
    ("scripts.modulo2.gen_06_papel_trabajo", "Módulo 2: Papel de Trabajo"),
    ("scripts.modulo2.gen_slides_m2", "Módulo 2: Slides"),
    ("scripts.modulo2.gen_pdf_ref_m2", "Módulo 2: PDF Referencia"),
    # Fase 4: Módulo 3
    ("scripts.modulo3.gen_07_combustible", "Módulo 3: Combustible"),
    ("scripts.modulo3.gen_08_comparativa_anual", "Módulo 3: Comparativa Anual"),
    ("scripts.modulo3.gen_slides_m3", "Módulo 3: Slides"),
    ("scripts.modulo3.gen_pdf_ref_m3", "Módulo 3: PDF Referencia"),
    # Fase 5: Módulo 4
    ("scripts.modulo4.gen_09_layout_dashboard", "Módulo 4: Layout Dashboard"),
    ("scripts.modulo4.gen_10_dashboard_final", "Módulo 4: Dashboard Final"),
    ("scripts.modulo4.gen_11_guia_proteccion", "Módulo 4: Guía Protección"),
    ("scripts.modulo4.gen_slides_m4", "Módulo 4: Slides"),
    ("scripts.modulo4.gen_pdf_ref_m4", "Módulo 4: PDF Referencia"),
    # Fase 6: Módulo 5
    ("scripts.modulo5.gen_12_dataset_copilot", "Módulo 5: Dataset Copilot"),
    ("scripts.modulo5.gen_prompts_copilot", "Módulo 5: Prompts Copilot"),
    ("scripts.modulo5.gen_slides_m5", "Módulo 5: Slides"),
    ("scripts.modulo5.gen_pdf_ref_m5", "Módulo 5: PDF Referencia"),
    # Fase 7: Bonus
    ("scripts.bonus.gen_slides_bonus1_vba", "Bonus 1: Slides VBA"),
    ("scripts.bonus.gen_slides_bonus2_claude_excel", "Bonus 2: Slides Claude"),
    ("scripts.bonus.gen_pdf_atajos", "Bonus: Cheat Sheet Atajos"),
    ("scripts.bonus.gen_pdf_bonus", "Bonus: PDFs (VBA + Claude)"),
]

# ── Archivos esperados ───────────────────────────────────────────
EXPECTED_FILES = {
    # Pack Excel Pro
    "Pack_Excel_Pro/Modulo_1_Funciones/01_Calculadora_ISR_V2026.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_1_Funciones/02_Control_Vencimientos_EFirma.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_1_Funciones/03_Extraccion_RFC_Master.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_1_Funciones/Referencia_Modulo_1.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Modulo_2_Tablas_Dinamicas/04_Limpieza_Masiva_Layout.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_2_Tablas_Dinamicas/05_Analisis_Nomina_XML_Pivot.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_2_Tablas_Dinamicas/06_Papel_Trabajo_Referenciado.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_2_Tablas_Dinamicas/Referencia_Modulo_2.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Modulo_3_Visualizacion/07_Dashboard_Ventas_Combustible.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_3_Visualizacion/08_Comparativa_Anual_Ventas_Gastos.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_3_Visualizacion/Referencia_Modulo_3.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Modulo_4_Dashboard/09_Layout_Dashboard_Contable.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_4_Dashboard/10_Dashboard_Final_Integrado.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_4_Dashboard/11_Guia_Proteccion_y_Seguridad.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Modulo_4_Dashboard/Referencia_Modulo_4.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Modulo_5_Copilot_IA/12_Dataset_Master_Copilot.xlsx": {"type": "xlsx"},
    "Pack_Excel_Pro/Modulo_5_Copilot_IA/Guia_Prompts_Copilot_Contadores.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Modulo_5_Copilot_IA/Referencia_Modulo_5.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Bonus/Guia_VBA_con_IA.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Bonus/Guia_Claude_en_Excel.pdf": {"type": "pdf"},
    "Pack_Excel_Pro/Bonus/Atajos_Excel_CheatSheet.pdf": {"type": "pdf"},
    # Slides
    "Slides/Modulo_1_Logica_Contable.pptx": {"type": "pptx"},
    "Slides/Modulo_2_Tablas_Dinamicas.pptx": {"type": "pptx"},
    "Slides/Modulo_3_Visualizacion.pptx": {"type": "pptx"},
    "Slides/Modulo_4_Dashboard.pptx": {"type": "pptx"},
    "Slides/Modulo_5_Copilot_IA.pptx": {"type": "pptx"},
    "Slides/Bonus_1_VBA_con_IA.pptx": {"type": "pptx"},
    "Slides/Bonus_2_Claude_en_Excel.pptx": {"type": "pptx"},
    # Teleprompter
    "Teleprompter/Script_Modulo_1.md": {"type": "md"},
    "Teleprompter/Script_Modulo_2.md": {"type": "md"},
    "Teleprompter/Script_Modulo_3.md": {"type": "md"},
    "Teleprompter/Script_Modulo_4.md": {"type": "md"},
    "Teleprompter/Script_Modulo_5.md": {"type": "md"},
    "Teleprompter/Script_Bonus_1_VBA.md": {"type": "md"},
    "Teleprompter/Script_Bonus_2_Claude.md": {"type": "md"},
}


def run_all():
    """Ejecuta todos los generadores."""
    print("=" * 60)
    print("  GENERANDO MATERIALES DEL CURSO")
    print("  Excel para Contadores y Administrativos")
    print("=" * 60)

    start = time.time()
    errors = []

    for module_path, description in GENERATORS:
        print(f"\n▶ {description}")
        try:
            mod = importlib.import_module(module_path)
            mod.build()
        except Exception as e:
            print(f"  ✗ ERROR: {e}")
            errors.append((description, str(e)))

    elapsed = time.time() - start
    print(f"\n{'=' * 60}")
    print(f"  Generación completada en {elapsed:.1f}s")

    if errors:
        print(f"  ⚠ {len(errors)} errores:")
        for desc, err in errors:
            print(f"    - {desc}: {err}")
    else:
        print("  ✓ Sin errores")

    print(f"{'=' * 60}")
    return len(errors) == 0


def verify():
    """Verifica que todos los archivos esperados existan y tengan contenido."""
    from scripts.config.constants import OUTPUT

    print("\n" + "=" * 60)
    print("  VERIFICACIÓN DE OUTPUTS")
    print("=" * 60)

    ok = 0
    missing = 0
    empty = 0

    for rel_path, meta in sorted(EXPECTED_FILES.items()):
        full_path = OUTPUT / rel_path
        if not full_path.exists():
            print(f"  ✗ FALTA: {rel_path}")
            missing += 1
        elif full_path.stat().st_size == 0:
            print(f"  ✗ VACÍO: {rel_path}")
            empty += 1
        else:
            size_kb = full_path.stat().st_size / 1024
            print(f"  ✓ {rel_path} ({size_kb:.1f} KB)")
            ok += 1

    total = ok + missing + empty
    print(f"\n{'=' * 60}")
    print(f"  Resultados: {ok}/{total} OK | {missing} faltantes | {empty} vacíos")

    # XLSX validation
    xlsx_count = 0
    xlsx_ok = 0
    try:
        from openpyxl import load_workbook
        for rel_path, meta in EXPECTED_FILES.items():
            if meta["type"] == "xlsx":
                full_path = OUTPUT / rel_path
                if full_path.exists():
                    xlsx_count += 1
                    try:
                        wb = load_workbook(str(full_path), read_only=True)
                        sheets = wb.sheetnames
                        wb.close()
                        if len(sheets) > 0:
                            xlsx_ok += 1
                        else:
                            print(f"  ⚠ {rel_path}: 0 hojas")
                    except Exception as e:
                        print(f"  ⚠ {rel_path}: Error al abrir: {e}")
        if xlsx_count > 0:
            print(f"  XLSX: {xlsx_ok}/{xlsx_count} abren correctamente")
    except ImportError:
        print("  (openpyxl no disponible para validación de XLSX)")

    # PPTX validation
    pptx_count = 0
    pptx_ok = 0
    try:
        from pptx import Presentation
        for rel_path, meta in EXPECTED_FILES.items():
            if meta["type"] == "pptx":
                full_path = OUTPUT / rel_path
                if full_path.exists():
                    pptx_count += 1
                    try:
                        prs = Presentation(str(full_path))
                        n_slides = len(prs.slides)
                        if n_slides >= 10:
                            pptx_ok += 1
                        else:
                            print(f"  ⚠ {rel_path}: solo {n_slides} slides")
                    except Exception as e:
                        print(f"  ⚠ {rel_path}: Error al abrir: {e}")
        if pptx_count > 0:
            print(f"  PPTX: {pptx_ok}/{pptx_count} tienen ≥10 slides")
    except ImportError:
        print("  (python-pptx no disponible para validación)")

    # MD validation
    md_count = 0
    md_ok = 0
    for rel_path, meta in EXPECTED_FILES.items():
        if meta["type"] == "md":
            full_path = OUTPUT / rel_path
            if full_path.exists():
                md_count += 1
                content = full_path.read_text(encoding="utf-8")
                sections = content.count("## ")
                words = len(content.split())
                if sections >= 5 and words >= 1000:
                    md_ok += 1
                else:
                    print(f"  ⚠ {rel_path}: {sections} secciones, {words} palabras")
    if md_count > 0:
        print(f"  MD Scripts: {md_ok}/{md_count} tienen ≥5 secciones y ≥1000 palabras")

    print(f"{'=' * 60}")
    return missing == 0 and empty == 0


if __name__ == "__main__":
    if "--verify" in sys.argv:
        verify()
    else:
        success = run_all()
        verify()
        sys.exit(0 if success else 1)
