"""
Generador: 06_Papel_Trabajo_Referenciado.xlsx
Modulo 2 -- Procesamiento Masivo y Analisis con Tablas Dinamicas

Hojas:
  - Papel_Trabajo: Formato de papel de trabajo para declaracion anual ISR
    Secciones: Ingresos, Deducciones, Base Gravable, Calculo ISR
    Usa IFERROR(VLOOKUP(...)) para buscar en la tarifa
  - Tarifa_Anual_2026: Tabla de tarifa ISR anual
  - Instrucciones: como vincular con tablas dinamicas
"""
import sys
from pathlib import Path
from typing import List, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from scripts.config.constants import PACK, Color, INSTRUCTOR, ANIO
from scripts.config.isr_2026 import TARIFA_ANUAL
from scripts.config.styles import (
    FILL_HEADER, FILL_LIGHT, FILL_VERDE, FILL_AMARILLO,
    FONT_HEADER, FONT_TITULO_XL, FONT_SUBTITULO, FONT_NORMAL, FONT_SMALL,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_RIGHT,
    FMT_MONEY, FMT_PCT,
    apply_header_style, apply_data_style, style_title_cell, auto_width
)
from scripts.generators.xlsx_gen import ExcelGenerator

OUTPUT_DIR = PACK / "Modulo_2_Tablas_Dinamicas"

# ── Estilos adicionales para el papel de trabajo ──────────────────
FILL_SECCION = PatternFill("solid", fgColor="DBEAFE")  # azul claro
FILL_RESULTADO = PatternFill("solid", fgColor="D1FAE5")  # verde claro
FILL_INPUT = PatternFill("solid", fgColor="FEF3C7")  # amarillo claro
FONT_SECCION = Font(name="Calibri", bold=True, size=12, color=Color.AZUL)
FONT_RESULTADO = Font(name="Calibri", bold=True, size=11, color="006100")
BORDER_BOTTOM_THICK = Border(
    bottom=Side(style="medium", color=Color.AZUL)
)


def _write_label_value(ws, row, label, value=None, is_formula=False,
                       is_input=False, is_result=False, explanation=""):
    """Escribe una fila de etiqueta + valor en el papel de trabajo."""
    cell_a = ws.cell(row=row, column=2, value=label)
    cell_a.font = FONT_NORMAL
    cell_a.border = THIN_BORDER
    cell_a.alignment = ALIGN_LEFT

    cell_b = ws.cell(row=row, column=3, value=value)
    cell_b.border = THIN_BORDER
    cell_b.alignment = ALIGN_RIGHT
    cell_b.number_format = FMT_MONEY

    if is_input:
        cell_b.fill = FILL_INPUT
        cell_b.font = FONT_SUBTITULO
    elif is_result:
        cell_b.fill = FILL_RESULTADO
        cell_b.font = FONT_RESULTADO
    else:
        cell_b.font = FONT_NORMAL

    if explanation:
        cell_c = ws.cell(row=row, column=4, value=explanation)
        cell_c.font = FONT_SMALL
        cell_c.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)


def _write_section_header(ws, row, title):
    """Escribe un encabezado de seccion."""
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=title)
    cell.font = FONT_SECCION
    cell.fill = FILL_SECCION
    cell.alignment = ALIGN_LEFT
    cell.border = BORDER_BOTTOM_THICK
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = FILL_SECCION
        ws.cell(row=row, column=col).border = BORDER_BOTTOM_THICK


def build():
    gen = ExcelGenerator("06_Papel_Trabajo_Referenciado.xlsx", OUTPUT_DIR)

    # ── Hoja 1: Tarifa Anual 2026 ──────────────────────────────
    ws_tarifa = gen.add_sheet("Tarifa_Anual_2026")
    style_title_cell(ws_tarifa, 1, 1,
                     "Tarifa ISR Anual 2026 -- Art. 152 LISR (Anexo 8 RMF)", 6)

    headers_t = ["Limite Inferior", "Limite Superior", "Cuota Fija",
                 "% Sobre Excedente"]
    data_t = []
    for r in TARIFA_ANUAL:
        sup = "En adelante" if r["lim_sup"] > 999_999_999 else r["lim_sup"]
        data_t.append([r["lim_inf"], sup, r["cuota"], r["pct"] / 100])

    gen.write_table(ws_tarifa, headers_t, data_t, start_row=3,
                    table_name="Tarifa_ISR_Anual",
                    money_cols=[1, 2, 3], pct_cols=[4])

    # ── Hoja 2: Papel de Trabajo ────────────────────────────────
    ws = gen.add_sheet("Papel_Trabajo")

    # Titulo principal
    ws.merge_cells("B2:D2")
    cell_titulo = ws.cell(row=2, column=2,
                          value="PAPEL DE TRABAJO - DECLARACION ANUAL ISR {}".format(ANIO))
    cell_titulo.font = FONT_TITULO_XL
    cell_titulo.alignment = ALIGN_CENTER

    ws.merge_cells("B3:D3")
    ws.cell(row=3, column=2,
            value="Contribuyente: [NOMBRE O RAZON SOCIAL]").font = FONT_SUBTITULO
    ws.cell(row=3, column=2).alignment = ALIGN_CENTER

    ws.merge_cells("B4:D4")
    ws.cell(row=4, column=2,
            value="RFC: [RFC DEL CONTRIBUYENTE]   |   Ejercicio: {}".format(ANIO)).font = FONT_SMALL
    ws.cell(row=4, column=2).alignment = ALIGN_CENTER

    r = 6  # fila actual

    # ── SECCION 1: INGRESOS ─────────────────────────────────────
    _write_section_header(ws, r, "I. INGRESOS ACUMULABLES")
    r += 1

    _write_label_value(ws, r, "Sueldos y salarios gravados", 0,
                       is_input=True,
                       explanation="Vincula con Pivot: Suma de ImporteGravado donde Clase=Percepcion")
    r += 1
    _write_label_value(ws, r, "Ingresos asimilados a salarios", 0,
                       is_input=True,
                       explanation="Si aplica, sumar concepto 046-Asimilados")
    r += 1
    _write_label_value(ws, r, "Otros ingresos acumulables", 0,
                       is_input=True,
                       explanation="Honorarios, arrendamiento, actividad empresarial, etc.")
    r += 1
    _write_label_value(ws, r, "TOTAL INGRESOS ACUMULABLES",
                       "=SUM(C7:C{})".format(r - 1),
                       is_result=True,
                       explanation="Suma de todos los ingresos")
    total_ingresos_row = r
    r += 2

    # ── SECCION 2: DEDUCCIONES ──────────────────────────────────
    _write_section_header(ws, r, "II. DEDUCCIONES PERSONALES (Art. 151 LISR)")
    r += 1

    deducciones_items = [
        ("Gastos medicos, dentales y hospitalarios", "Maximo 15% de ingresos o 5 UMAs anuales"),
        ("Gastos funerarios", "Hasta 1 UMA anual ($41,297)"),
        ("Donativos", "Hasta 7% de ingresos acumulables del ejercicio anterior"),
        ("Intereses reales hipotecarios", "Creditos contratados con SOFOMES/bancos"),
        ("Aportaciones voluntarias al retiro", "Hasta 10% de ingresos o 5 UMAs anuales"),
        ("Primas por seguros de gastos medicos", "Pagos complementarios al IMSS"),
        ("Transporte escolar obligatorio", "Solo si es obligatorio por la escuela"),
        ("Colegiaturas (estimulo fiscal)", "Preescolar $14,200; Primaria $12,900; Secundaria $19,900"),
    ]

    first_ded_row = r
    for label, expl in deducciones_items:
        _write_label_value(ws, r, label, 0, is_input=True, explanation=expl)
        r += 1
    last_ded_row = r - 1

    _write_label_value(ws, r, "TOTAL DEDUCCIONES PERSONALES",
                       "=SUM(C{}:C{})".format(first_ded_row, last_ded_row),
                       is_result=True,
                       explanation="Tope: 15% de ingresos o 5 UMAs anuales (lo menor)")
    total_deducciones_row = r
    r += 2

    # ── SECCION 3: BASE GRAVABLE ────────────────────────────────
    _write_section_header(ws, r, "III. BASE GRAVABLE")
    r += 1

    _write_label_value(ws, r, "Total Ingresos Acumulables",
                       "=C{}".format(total_ingresos_row),
                       explanation="Referencia a seccion I")
    r += 1
    _write_label_value(ws, r, "(-) Total Deducciones Personales",
                       "=C{}".format(total_deducciones_row),
                       explanation="Referencia a seccion II")
    r += 1
    _write_label_value(ws, r, "BASE GRAVABLE DEL EJERCICIO",
                       "=MAX(0,C{}-C{})".format(r - 2, r - 1),
                       is_result=True,
                       explanation="No puede ser negativa")
    base_gravable_row = r
    r += 2

    # ── SECCION 4: CALCULO ISR ──────────────────────────────────
    _write_section_header(ws, r, "IV. CALCULO DEL ISR (Art. 152 LISR)")
    r += 1

    bg_ref = "C{}".format(base_gravable_row)

    _write_label_value(ws, r, "Base Gravable",
                       "={}".format(bg_ref),
                       explanation="De seccion III")
    bg_calc_row = r
    r += 1

    # Limite inferior con IFERROR(VLOOKUP(...))
    _write_label_value(ws, r, "Limite Inferior (BUSCARV)",
                       '=IFERROR(VLOOKUP(C{},Tarifa_ISR_Anual,1,TRUE),0)'.format(bg_calc_row),
                       explanation="BUSCARV aproximado en la tarifa")
    lim_inf_row = r
    r += 1

    _write_label_value(ws, r, "Excedente sobre Limite Inferior",
                       "=C{}-C{}".format(bg_calc_row, lim_inf_row),
                       explanation="Base Gravable - Limite Inferior")
    excedente_row = r
    r += 1

    _write_label_value(ws, r, "% Sobre Excedente (BUSCARV)",
                       '=IFERROR(VLOOKUP(C{},Tarifa_ISR_Anual,4,TRUE),0)'.format(bg_calc_row),
                       explanation="Porcentaje marginal de la tarifa")
    ws.cell(row=r, column=3).number_format = FMT_PCT
    pct_row = r
    r += 1

    _write_label_value(ws, r, "ISR Marginal",
                       "=C{}*C{}".format(excedente_row, pct_row),
                       explanation="Excedente x Porcentaje")
    isr_marginal_row = r
    r += 1

    _write_label_value(ws, r, "Cuota Fija (BUSCARV)",
                       '=IFERROR(VLOOKUP(C{},Tarifa_ISR_Anual,3,TRUE),0)'.format(bg_calc_row),
                       explanation="Cuota fija de la tarifa")
    cuota_fija_row = r
    r += 1

    _write_label_value(ws, r, "ISR CAUSADO DEL EJERCICIO",
                       "=C{}+C{}".format(isr_marginal_row, cuota_fija_row),
                       is_result=True,
                       explanation="ISR Marginal + Cuota Fija")
    isr_causado_row = r
    r += 2

    # ── SECCION 5: ISR A CARGO / FAVOR ──────────────────────────
    _write_section_header(ws, r, "V. DETERMINACION DEL ISR A CARGO O A FAVOR")
    r += 1

    _write_label_value(ws, r, "ISR Causado del Ejercicio",
                       "=C{}".format(isr_causado_row),
                       explanation="De seccion IV")
    isr_causado_ref = r
    r += 1

    _write_label_value(ws, r, "(-) Retenciones de ISR en el ejercicio", 0,
                       is_input=True,
                       explanation="Vincula con Pivot: Suma de Deduccion ISR del ejercicio")
    retenciones_row = r
    r += 1

    _write_label_value(ws, r, "(-) Pagos provisionales efectuados", 0,
                       is_input=True,
                       explanation="Solo si realizaste pagos provisionales (Cap. II, III, IV LISR)")
    pagos_prov_row = r
    r += 1

    _write_label_value(ws, r, "ISR A CARGO (+) O A FAVOR (-)",
                       "=C{}-C{}-C{}".format(isr_causado_ref, retenciones_row, pagos_prov_row),
                       is_result=True,
                       explanation="Positivo = a cargo; Negativo = a favor (solicitar devolucion)")
    r += 2

    # ── Nota al pie ─────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(row=r, column=2,
            value="Nota: Las celdas amarillas son de captura manual. "
                  "Las verdes se calculan automaticamente con formulas.").font = FONT_SMALL
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(row=r, column=2,
            value="Las formulas BUSCARV buscan en la tabla 'Tarifa_ISR_Anual' de la hoja Tarifa_Anual_2026.").font = FONT_SMALL
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(row=r, column=2,
            value="Preparado por: {} | {} | Tarifa vigente Anexo 8 RMF {}".format(
                INSTRUCTOR, ANIO, ANIO)).font = FONT_SMALL

    # Ajustar anchos
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 55

    # ── Hoja 3: Instrucciones ───────────────────────────────────
    gen.add_instructions_sheet([
        "Este papel de trabajo calcula el ISR anual usando la tarifa 2026 (Art. 152 LISR).",
        "Las celdas AMARILLAS son de captura: ingresa ahi los montos de tus ingresos y deducciones.",
        "Las celdas VERDES se calculan automaticamente con formulas.",
        "",
        "=== VINCULACION CON TABLAS DINAMICAS ===",
        "1. Abre el archivo 05_Analisis_Nomina_XML_Pivot.xlsx.",
        "2. Crea una Tabla Dinamica que sume ImporteGravado por Clase.",
        "3. Copia el total de 'Percepcion' en la celda 'Sueldos y salarios gravados' (C7).",
        "4. Copia el total de 'Deduccion' con concepto ISR en 'Retenciones de ISR' (seccion V).",
        "5. Observa como el papel de trabajo calcula automaticamente el ISR a cargo o a favor.",
        "",
        "=== FORMULAS CLAVE ===",
        "BUSCARV (VLOOKUP): Busca el limite inferior, cuota fija y porcentaje en la tarifa.",
        "SI.ERROR (IFERROR): Envuelve BUSCARV para evitar errores si la celda esta vacia.",
        "MAX(0, valor): Asegura que la base gravable no sea negativa.",
        "",
        "=== PARA EL INSTRUCTOR ===",
        "Este formato replica el calculo que hace el SAT en la declaracion anual.",
        "Es ideal para mostrar la relacion entre tablas dinamicas y papeles de trabajo.",
        "Los alumnos pueden verificar sus resultados contra el simulador del SAT.",
    ])

    gen.save()


if __name__ == "__main__":
    build()
