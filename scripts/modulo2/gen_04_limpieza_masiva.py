"""
Generador: 04_Limpieza_Masiva_Layout.xlsx
Modulo 2 -- Procesamiento Masivo y Analisis con Tablas Dinamicas

Hojas:
  - Datos_Sucios: 200+ filas con problemas intencionales
    (celdas vacias, formatos inconsistentes, texto como numero,
     espacios extra, fechas mixtas)
  - Datos_Limpios: referencia de como deben lucir los datos
  - Instrucciones
"""
import sys
from pathlib import Path
import random
from datetime import date, timedelta
from typing import List, Tuple, Optional

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl.styles import PatternFill, Font, Alignment, numbers
from scripts.config.constants import PACK
from scripts.config.styles import (
    FILL_AMARILLO, FILL_VERDE, FILL_ROJO, FILL_LIGHT,
    FONT_HEADER, FONT_NORMAL, FONT_SMALL, FONT_SUBTITULO,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_RIGHT,
    FMT_MONEY, FMT_DATE,
    apply_header_style, apply_data_style, style_title_cell, auto_width
)
from scripts.generators.xlsx_gen import ExcelGenerator

OUTPUT_DIR = PACK / "Modulo_2_Tablas_Dinamicas"

# ── Datos maestros para generacion ───────────────────────────────
PROVEEDORES = [
    ("Distribuidora del Norte SA de CV", "DNO850315XX1"),
    ("Grupo Papelero Nacional SA", "GPN920801KL3"),
    ("Combustibles y Lubricantes Azteca SA de CV", "CLA100512M45"),
    ("Servicios de Limpieza Integral SC", "SLI880620PQ2"),
    ("Tecnologia Avanzada del Bajio SA", "TAB760910RS7"),
    ("Consultores en TI Hernandez SC", "CTH051103TU8"),
    ("Transportes y Mensajeria Reyes SA de CV", "TMR110215VW9"),
    ("Material Electrico Santa Cruz SA de CV", "MES990430XY0"),
    ("Constructora Inmobiliaria Vega SA", "CIV070817ZA1"),
    ("Suministros Medicos Martinez SC", "SMM130225BC4"),
    ("Alimentos y Bebidas del Sureste SA", "ABS170610CD5"),
    ("Imprenta Digital Express SC", "IDE200315EF6"),
    ("Seguridad Privada Nacional SA de CV", "SPN080923GH7"),
    ("Refacciones Automotrices Lopez SA", "RAL150412IJ8"),
    ("Papeleria y Oficina Central SC", "POC190801KL9"),
]

CONCEPTOS = [
    "Papeleria y articulos de oficina",
    "Combustible vehiculos utilitarios",
    "Servicio de limpieza mensual",
    "Mantenimiento equipo de computo",
    "Renta de oficina mensual",
    "Suministros medicos",
    "Material electrico",
    "Servicio de transporte",
    "Impresion de documentos",
    "Servicio de seguridad",
    "Alimentos para eventos",
    "Refacciones automotrices",
    "Licencias de software",
    "Capacitacion al personal",
    "Mobiliario de oficina",
]

# Formatos de fecha "sucios" para simular inconsistencias
DATE_FORMATS_DIRTY = [
    lambda d: d.strftime("%d/%m/%Y"),        # 15/03/2025
    lambda d: d.strftime("%d-%m-%Y"),         # 15-03-2025
    lambda d: d.strftime("%Y-%m-%d"),         # 2025-03-15
    lambda d: d.strftime("%d %b %Y"),         # 15 Mar 2025
    lambda d: d.strftime("%m/%d/%Y"),         # 03/15/2025
    lambda d: d.strftime("%d.%m.%Y"),         # 15.03.2025
    lambda d: d,                              # date object (clean)
]


def _gen_dirty_row(idx, rng):
    # type: (int, random.Random) -> Tuple[List, List]
    """Genera una fila sucia y su version limpia correspondiente."""
    prov_name, prov_rfc = rng.choice(PROVEEDORES)
    concepto = rng.choice(CONCEPTOS)
    base_date = date(2025, 1, 1) + timedelta(days=rng.randint(0, 364))
    subtotal = round(rng.uniform(500, 85000), 2)
    iva = round(subtotal * 0.16, 2)
    total = round(subtotal + iva, 2)
    folio = "F-{:04d}".format(idx)

    # ── Version limpia ──
    clean_row = [
        folio, prov_name, prov_rfc, base_date,
        concepto, subtotal, iva, total, ""
    ]

    # ── Version sucia: introducir problemas aleatorios ──
    dirty_folio = folio
    dirty_prov = prov_name
    dirty_rfc = prov_rfc
    dirty_date = base_date
    dirty_concepto = concepto
    dirty_subtotal = subtotal
    dirty_iva = iva
    dirty_total = total
    dirty_obs = ""

    problem = rng.random()

    # 12% prob: celda vacia en Proveedor
    if problem < 0.12:
        dirty_prov = None
        dirty_obs = "Proveedor faltante"

    # 10% prob: RFC con espacio extra
    elif problem < 0.22:
        dirty_rfc = "  " + prov_rfc + " "
        dirty_obs = "RFC con espacios"

    # 10% prob: subtotal como texto
    elif problem < 0.32:
        dirty_subtotal = "$" + "{:,.2f}".format(subtotal)
        dirty_obs = "Subtotal como texto"

    # 8% prob: fecha en formato texto (string)
    elif problem < 0.40:
        fmt_fn = rng.choice(DATE_FORMATS_DIRTY[:6])
        dirty_date = fmt_fn(base_date)
        dirty_obs = "Fecha como texto"

    # 8% prob: IVA vacio
    elif problem < 0.48:
        dirty_iva = None
        dirty_total = None
        dirty_obs = "IVA y Total faltantes"

    # 7% prob: concepto con espacios extra y mayusculas inconsistentes
    elif problem < 0.55:
        dirty_concepto = "  " + concepto.upper() + "   "
        dirty_obs = "Concepto con espacios y mayusculas"

    # 7% prob: folio duplicado (se usara el anterior)
    elif problem < 0.62:
        dirty_folio = "F-{:04d}".format(max(1, idx - 1))
        dirty_obs = "Folio duplicado"

    # 6% prob: total no cuadra (subtotal + iva != total)
    elif problem < 0.68:
        dirty_total = round(total + rng.uniform(10, 500), 2)
        dirty_obs = "Total no cuadra"

    # 6% prob: nombre proveedor con errores tipograficos
    elif problem < 0.74:
        dirty_prov = prov_name.replace("SA de CV", "S.A. de C.V.").replace("SC", "S.C.")
        dirty_obs = "Razon social inconsistente"

    # 5% prob: RFC invalido (muy corto)
    elif problem < 0.79:
        dirty_rfc = prov_rfc[:8]
        dirty_obs = "RFC incompleto"

    # 5% prob: numeros negativos
    elif problem < 0.84:
        dirty_subtotal = -subtotal
        dirty_iva = -iva
        dirty_total = -total
        dirty_obs = "Montos negativos"

    # 5% prob: factura cancelada — fila irrecuperable, se debe eliminar
    elif problem < 0.89:
        dirty_prov = None
        dirty_rfc = None
        dirty_date = None
        dirty_concepto = None
        dirty_subtotal = None
        dirty_iva = None
        dirty_total = None
        dirty_obs = "Factura cancelada - sin datos"
        clean_row = None  # No debe aparecer en Datos_Limpios

    # El resto: datos correctos (sin errores)
    # ~11% quedan limpios

    dirty_row = [
        dirty_folio, dirty_prov, dirty_rfc, dirty_date,
        dirty_concepto, dirty_subtotal, dirty_iva, dirty_total, dirty_obs
    ]

    return dirty_row, clean_row


def build():
    gen = ExcelGenerator("04_Limpieza_Masiva_Layout.xlsx", OUTPUT_DIR)

    HEADERS = [
        "Folio", "Proveedor", "RFC", "Fecha",
        "Concepto", "Subtotal", "IVA", "Total", "Observaciones"
    ]

    rng = random.Random(42)
    dirty_rows = []
    clean_rows = []

    for i in range(1, 221):
        dirty, clean = _gen_dirty_row(i, rng)
        dirty_rows.append(dirty)
        if clean is not None:
            clean_rows.append(clean)

    # ── Hoja 1: Datos Sucios ────────────────────────────────────
    ws1 = gen.add_sheet("Datos_Sucios")
    style_title_cell(ws1, 1, 1,
                     "Datos de Compras 2025 - PENDIENTE DE LIMPIEZA", 9)
    ws1.cell(row=2, column=1,
             value="Estos datos tienen errores intencionales. Tu mision: limpiarlos.").font = FONT_SMALL

    # Write headers manually (not as table -- messy data shouldn't be a table yet)
    for c, h in enumerate(HEADERS, 1):
        cell = ws1.cell(row=4, column=c, value=h)
    apply_header_style(ws1, 4, 1, len(HEADERS))

    # Write dirty data WITHOUT uniform styling to emphasize messiness
    for r_idx, row_data in enumerate(dirty_rows):
        row_num = 5 + r_idx
        for c_idx, val in enumerate(row_data):
            cell = ws1.cell(row=row_num, column=c_idx + 1, value=val)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER

        # If Observaciones has content, highlight row lightly in yellow
        if row_data[8]:
            for c in range(1, len(HEADERS) + 1):
                ws1.cell(row=row_num, column=c).fill = PatternFill(
                    "solid", fgColor="FFF3CD"
                )

    # Set reasonable column widths
    widths = {"A": 10, "B": 42, "C": 16, "D": 16, "E": 38,
              "F": 14, "G": 14, "H": 14, "I": 32}
    for col_letter, w in widths.items():
        ws1.column_dimensions[col_letter].width = w

    # ── Hoja 2: Datos Limpios (referencia) ──────────────────────
    ws2 = gen.add_sheet("Datos_Limpios")
    style_title_cell(ws2, 1, 1,
                     "Datos de Compras 2025 - REFERENCIA LIMPIA", 9)
    ws2.cell(row=2, column=1,
             value="Asi deben lucir los datos despues de la limpieza.").font = FONT_SMALL

    gen.write_table(ws2, HEADERS, clean_rows, start_row=4,
                    table_name="compras_limpias",
                    money_cols=[6, 7, 8], date_cols=[4])

    # ── Hoja 3: Instrucciones ───────────────────────────────────
    gen.add_instructions_sheet([
        "La hoja 'Datos_Sucios' contiene 220 filas con errores intencionales que debes limpiar.",
        "Tipos de errores incluidos: celdas vacias, espacios extra, texto donde deberia haber numeros, "
        "formatos de fecha inconsistentes, RFCs incompletos, totales que no cuadran y mas.",
        "Las filas con fondo amarillo claro tienen errores detectados en la columna 'Observaciones'.",
        "PASO 1: Usa Buscar y Reemplazar (Ctrl+H) para quitar signos '$' y comas en montos.",
        "PASO 2: Usa ESPACIOS (TRIM) para eliminar espacios extra en RFC y nombres.",
        "PASO 3: Verifica que Subtotal + IVA = Total usando una columna auxiliar.",
        "PASO 4: Estandariza las fechas al formato DD/MM/AAAA con DATEVALUE o pegado especial.",
        "PASO 5: Elimina las filas de facturas canceladas (solo tienen folio, sin datos recuperables).",
        "PASO 6: Quita duplicados con la funcion Datos > Quitar duplicados.",
        "PASO 7: Cuando los datos esten limpios, convierte el rango a Tabla (Ctrl+T).",
        "La hoja 'Datos_Limpios' es tu referencia: compara tu resultado contra ella.",
        "RETO: Intenta hacer la limpieza en menos de 15 minutos usando atajos de Excel.",
    ])

    gen.save()


if __name__ == "__main__":
    build()
