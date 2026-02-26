"""
Generador: 09_Layout_Dashboard_Contable.xlsx
Modulo 4 -- El Dashboard Inteligente y Entrega Profesional

Template con areas designadas para construir un dashboard contable:
  - Area de KPIs (filas 1-5): 4 cuadros de indicadores clave
  - Area de graficos (filas 6-25): placeholders para Chart 1 y Chart 2
  - Area de filtros (columnas A-B): zona para segmentadores
  - Hoja de instrucciones
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl.utils import get_column_letter
from scripts.config.constants import PACK, Color
from scripts.config.styles import (
    FILL_HEADER, FILL_LIGHT, FILL_MEDIO, FILL_VERDE, FILL_AMARILLO, FILL_BLANCO,
    FONT_HEADER, FONT_TITULO_XL, FONT_SUBTITULO, FONT_NORMAL, FONT_SMALL,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_RIGHT,
    FMT_MONEY, style_title_cell, auto_width,
    PatternFill, Font, Alignment, Border, Side
)
from scripts.generators.xlsx_gen import ExcelGenerator

OUTPUT_DIR = PACK / "Modulo_4_Dashboard"


def _make_kpi_box(ws, row_start, col_start, title, value_text, fill_color, font_color):
    """Create a styled KPI box spanning 4 rows x 3 columns."""
    col_end = col_start + 2

    # Merge title row
    ws.merge_cells(
        start_row=row_start, start_column=col_start,
        end_row=row_start, end_column=col_end
    )
    title_cell = ws.cell(row=row_start, column=col_start, value=title)
    title_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=fill_color)
    title_cell.alignment = ALIGN_CENTER
    title_cell.border = THIN_BORDER

    # Fill remaining header row cells border
    for c in range(col_start + 1, col_end + 1):
        cell = ws.cell(row=row_start, column=c)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.border = THIN_BORDER

    # Merge value row
    ws.merge_cells(
        start_row=row_start + 1, start_column=col_start,
        end_row=row_start + 3, end_column=col_end
    )
    val_cell = ws.cell(row=row_start + 1, column=col_start, value=value_text)
    val_cell.font = Font(name="Calibri", bold=True, size=16, color=font_color)
    val_cell.alignment = ALIGN_CENTER
    val_cell.border = THIN_BORDER

    # Apply borders to all cells in the box
    for r in range(row_start, row_start + 4):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def _make_placeholder(ws, row_start, row_end, col_start, col_end, label):
    """Create a placeholder area with a centered label and light border."""
    # Merge center area for label
    mid_row = (row_start + row_end) // 2
    ws.merge_cells(
        start_row=mid_row, start_column=col_start,
        end_row=mid_row, end_column=col_end
    )
    cell = ws.cell(row=mid_row, column=col_start, value=label)
    cell.font = Font(name="Calibri", bold=False, size=14, color=Color.TEXTO_MEDIO)
    cell.alignment = ALIGN_CENTER

    # Light fill for entire area
    light_fill = PatternFill("solid", fgColor=Color.FONDO_MEDIO)
    border_dashed = Border(
        left=Side(style="dashed", color=Color.GRIS_BORDE),
        right=Side(style="dashed", color=Color.GRIS_BORDE),
        top=Side(style="dashed", color=Color.GRIS_BORDE),
        bottom=Side(style="dashed", color=Color.GRIS_BORDE),
    )
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = light_fill
            cell.border = border_dashed


def build():
    gen = ExcelGenerator("09_Layout_Dashboard_Contable.xlsx", OUTPUT_DIR)

    # ---- Hoja 1: Dashboard Layout ----------------------------------------
    ws = gen.add_sheet("Dashboard_Layout")

    # Background fill for entire visible area
    bg_fill = PatternFill("solid", fgColor=Color.FONDO_CLARO)
    for r in range(1, 30):
        for c in range(1, 18):
            ws.cell(row=r, column=c).fill = bg_fill

    # Title row
    ws.merge_cells("A1:Q1")
    title_cell = ws.cell(row=1, column=1, value="PLANTILLA: Construye Tu Dashboard Aqui")
    title_cell.font = Font(name="Calibri", bold=True, size=18, color=Color.AZUL)
    title_cell.alignment = ALIGN_CENTER
    title_cell.fill = PatternFill("solid", fgColor=Color.BLANCO)
    ws.row_dimensions[1].height = 40

    # ---- KPI Area (rows 2-5) ----
    # KPI 1: Total Percepciones
    _make_kpi_box(ws, 2, 3, "Total Percepciones", "$0.00", Color.AZUL, Color.AZUL)
    # KPI 2: Total Deducciones
    _make_kpi_box(ws, 2, 7, "Total Deducciones", "$0.00", Color.ROJO, Color.ROJO)
    # KPI 3: ISR del Periodo
    _make_kpi_box(ws, 2, 11, "ISR del Periodo", "$0.00", Color.AMARILLO, Color.AMARILLO)
    # KPI 4: e.firma Status
    _make_kpi_box(ws, 2, 15, "e.firma Status", "VIGENTE", Color.VERDE, Color.VERDE)

    # Row heights for KPI area
    for r in range(2, 6):
        ws.row_dimensions[r].height = 25

    # ---- Filters area (column A-B, rows 7-25) ----
    ws.merge_cells("A7:B7")
    filt_title = ws.cell(row=7, column=1, value="FILTROS / SEGMENTADORES")
    filt_title.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    filt_title.fill = PatternFill("solid", fgColor=Color.AZUL)
    filt_title.alignment = ALIGN_CENTER
    ws.cell(row=7, column=2).fill = PatternFill("solid", fgColor=Color.AZUL)

    filter_labels = [
        "Periodo:", "Departamento:", "Puesto:", "Rango salarial:"
    ]
    for i, label in enumerate(filter_labels):
        r = 9 + i * 3
        ws.cell(row=r, column=1, value=label).font = FONT_NORMAL
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        # Placeholder for slicer
        ws.cell(row=r + 1, column=1, value="[Insertar segmentador aqui]").font = FONT_SMALL
        ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=2)

    # Set filter column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    # ---- Charts area (rows 7-25) ----
    # Chart 1: left chart area
    _make_placeholder(ws, 7, 16, 3, 9, "Grafico 1 aqui")
    # Chart 2: right chart area
    _make_placeholder(ws, 7, 16, 11, 17, "Grafico 2 aqui")
    # Chart 3: bottom area
    _make_placeholder(ws, 18, 25, 3, 17, "Grafico 3 aqui (tabla detalle o grafico adicional)")

    # Column widths for chart area
    for c in range(3, 18):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Footer
    ws.merge_cells("A27:Q27")
    footer = ws.cell(row=27, column=1,
                     value="Curso: Excel para Contadores y Administrativos | Israel Castro | 2026")
    footer.font = FONT_SMALL
    footer.alignment = ALIGN_CENTER

    # Print setup
    ws.sheet_properties.tabColor = Color.AZUL

    # ---- Hoja 2: Instrucciones ----
    gen.add_instructions_sheet([
        "PLANTILLA DE TRABAJO -- Usa este archivo para construir tu dashboard durante la clase.",
        "PASO 1: Abre los archivos de ejercicios anteriores (Modulos 1-3) donde tienes tus datos de nomina.",
        "PASO 2: Crea una Tabla Dinamica desde tus datos (Insertar > Tabla Dinamica) y pegala en esta hoja.",
        "PASO 3: Inserta Segmentadores vinculados a tu TD (clic en TD > Insertar > Segmentacion de datos).",
        "PASO 4: Crea graficos desde tu TD y colocalos en las areas marcadas con bordes punteados.",
        "PASO 5: Reemplaza '$0.00' en los KPIs con formulas =SUBTOTAL(109,...) que apunten a tu tabla.",
        "PASO 6: Oculta las lineas de cuadricula (Vista > desmarcar 'Lineas de cuadricula').",
        "Si quieres ver una solucion completa de referencia, abre el archivo: 10_Dashboard_Final_Integrado.xlsx",
        "Tip: Inmoviliza paneles en fila 5 para que los KPIs queden fijos al desplazarte.",
        "Este archivo es TU espacio de trabajo. El instructor trabaja en paralelo con el mismo template.",
    ])

    gen.save()


if __name__ == "__main__":
    build()
