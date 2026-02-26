"""
Generador: 10_Dashboard_Final_Integrado.xlsx
Modulo 4 -- El Dashboard Inteligente y Entrega Profesional

Hojas:
  - Datos_Nomina: 20 empleados x 12 meses con datos de nomina (tabla nombrada)
  - Tarifa_ISR: Tarifa mensual ISR 2026 completa (tabla nombrada)
  - Calculadora: Area para calculo ISR con VLOOKUP
  - Dashboard: Layout con KPIs, instrucciones para slicers y graficos
  - Instrucciones
"""
import sys
from pathlib import Path
import random

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl.utils import get_column_letter
from scripts.config.constants import PACK, Color, SALARIO_MINIMO_MENSUAL
from scripts.config.isr_2026 import calcular_isr_mensual
from scripts.config.styles import (
    FILL_VERDE, FILL_AMARILLO,
    FONT_SUBTITULO, FONT_NORMAL, FONT_SMALL,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT,
    FMT_MONEY,
    style_title_cell,
    PatternFill, Font, Alignment, Border, Side
)
from scripts.generators.xlsx_gen import ExcelGenerator

OUTPUT_DIR = PACK / "Modulo_4_Dashboard"


def _make_placeholder(ws, row_start, row_end, col_start, col_end, label):
    """Create a placeholder area with a centered label and dashed border."""
    mid_row = (row_start + row_end) // 2
    ws.merge_cells(
        start_row=mid_row, start_column=col_start,
        end_row=mid_row, end_column=col_end
    )
    cell = ws.cell(row=mid_row, column=col_start, value=label)
    cell.font = Font(name="Calibri", bold=False, size=14, color=Color.TEXTO_MEDIO)
    cell.alignment = ALIGN_CENTER

    light_fill = PatternFill("solid", fgColor=Color.FONDO_MEDIO)
    border_dashed = Border(
        left=Side(style="dashed", color=Color.GRIS_BORDE),
        right=Side(style="dashed", color=Color.GRIS_BORDE),
        top=Side(style="dashed", color=Color.GRIS_BORDE),
        bottom=Side(style="dashed", color=Color.GRIS_BORDE),
    )
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            c_cell = ws.cell(row=r, column=c)
            c_cell.fill = light_fill
            c_cell.border = border_dashed


# ---- Datos de empleados ----
EMPLEADOS = [
    ("Ana Lopez Martinez", "Contador Senior"),
    ("Carlos Ramirez Ortiz", "Auxiliar Contable"),
    ("Maria Fernanda Torres", "Gerente Administrativo"),
    ("Jorge Hernandez Cruz", "Analista Fiscal"),
    ("Patricia Gonzalez Ruiz", "Asistente Administrativo"),
    ("Roberto Sanchez Diaz", "Director Financiero"),
    ("Laura Garcia Mendoza", "Contador Junior"),
    ("Fernando Morales Rios", "Auditor Interno"),
    ("Gabriela Flores Castillo", "Nominas y IMSS"),
    ("Miguel Angel Vargas", "Gerente de Operaciones"),
    ("Diana Castro Perez", "Auxiliar de Nominas"),
    ("Alejandro Reyes Luna", "Contador General"),
    ("Sofia Martinez Aguilar", "Analista de Costos"),
    ("Ricardo Jimenez Torres", "Jefe de Contabilidad"),
    ("Claudia Romero Navarro", "Asistente Fiscal"),
    ("Eduardo Gutierrez Soto", "Tesorero"),
    ("Veronica Diaz Herrera", "Coordinadora RH"),
    ("Oscar Mendez Rojas", "Pasante Contable"),
    ("Isabel Juarez Medina", "Ejecutiva de Cobranza"),
    ("Raul Dominguez Espino", "Gerente General"),
]

PUESTOS_SALARIOS = {
    "Pasante Contable": (SALARIO_MINIMO_MENSUAL, 10000),
    "Auxiliar Contable": (10000, 14000),
    "Auxiliar de Nominas": (9500, 13000),
    "Asistente Administrativo": (10000, 14000),
    "Asistente Fiscal": (10500, 14500),
    "Contador Junior": (14000, 20000),
    "Contador Senior": (20000, 30000),
    "Contador General": (25000, 35000),
    "Analista Fiscal": (18000, 26000),
    "Analista de Costos": (17000, 25000),
    "Nominas y IMSS": (15000, 22000),
    "Jefe de Contabilidad": (28000, 38000),
    "Auditor Interno": (22000, 32000),
    "Coordinadora RH": (20000, 28000),
    "Ejecutiva de Cobranza": (12000, 18000),
    "Tesorero": (25000, 35000),
    "Gerente Administrativo": (32000, 42000),
    "Gerente de Operaciones": (35000, 45000),
    "Director Financiero": (40000, 55000),
    "Gerente General": (45000, 55000),
}

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]


def _calcular_imss(sueldo):
    """Cuota IMSS trabajador simplificada (~2.775% del SBC tope 25 UMA)."""
    uma_diario = 113.14  # UMA 2026 estimado
    tope_sbc = uma_diario * 25 * 30
    base = min(sueldo, tope_sbc)
    return round(base * 0.02775, 2)


def _calcular_subsidio(sueldo):
    """Subsidio al empleo simplificado para sueldos bajos."""
    from scripts.config.isr_2026 import SUBSIDIO_EMPLEO_MENSUAL
    for rango in SUBSIDIO_EMPLEO_MENSUAL:
        if rango["desde"] <= sueldo <= rango["hasta"]:
            return rango["subsidio"]
    return 0.0


def build():
    gen = ExcelGenerator("10_Dashboard_Final_Integrado.xlsx", OUTPUT_DIR)

    random.seed(2026)

    # ==== Hoja 1: Datos_Nomina ============================================
    ws1 = gen.add_sheet("Datos_Nomina")
    style_title_cell(ws1, 1, 1, "Datos de Nomina 2026 -- 20 Empleados x 12 Meses", 8)
    ws1.cell(row=2, column=1,
             value="Base de datos para Tablas Dinamicas, Segmentadores y Dashboard.").font = FONT_SMALL
    ws1.cell(row=3, column=1,
             value="Datos de ejemplo ya incluidos. En tu version real, reemplaza con tu propia tabla de nomina.").font = FONT_SMALL

    headers = [
        "Empleado", "Puesto", "Periodo", "Sueldo",
        "ISR", "IMSS", "SubsidioEmpleo", "NetoPagar"
    ]

    data = []
    for nombre, puesto in EMPLEADOS:
        sal_min, sal_max = PUESTOS_SALARIOS.get(puesto, (15000, 25000))
        sueldo_base = round(random.uniform(sal_min, sal_max), 2)

        for mes in MESES:
            # Small monthly variation (+/- 3% for bonuses/overtime)
            variacion = random.uniform(-0.03, 0.03)
            sueldo = round(sueldo_base * (1 + variacion), 2)

            isr_result = calcular_isr_mensual(sueldo)
            isr = isr_result.get("isr_total", 0.0) if isr_result else 0.0
            imss = _calcular_imss(sueldo)
            subsidio = _calcular_subsidio(sueldo)

            neto = round(sueldo - isr - imss + subsidio, 2)
            if neto < 0:
                neto = round(sueldo * 0.70, 2)

            data.append([nombre, puesto, mes, sueldo, isr, imss, subsidio, neto])

    gen.write_table(
        ws1, headers, data, start_row=4,
        table_name="Nomina_Empleados",
        money_cols=[4, 5, 6, 7, 8]
    )

    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 26
    ws1.column_dimensions["C"].width = 14

    # ==== Hoja 2: Dashboard ===============================================
    ws2 = gen.add_sheet("Dashboard")
    style_title_cell(ws2, 1, 1, "SOLUCION COMPLETA: Dashboard de Nomina Integrado", 12)

    # Background
    bg_fill = PatternFill("solid", fgColor=Color.FONDO_CLARO)
    for r in range(1, 30):
        for c in range(1, 14):
            ws2.cell(row=r, column=c).fill = bg_fill

    ws2.cell(row=1, column=1).fill = PatternFill("solid", fgColor=Color.BLANCO)
    ws2.row_dimensions[1].height = 35

    # Context banner (row 2)
    ws2.merge_cells("A2:M2")
    banner4 = ws2.cell(
        row=2, column=1,
        value="Ejemplo terminado con datos reales | Para construir el tuyo desde cero: 09_Layout_Dashboard_Contable.xlsx"
    )
    banner4.font = FONT_SMALL
    banner4.fill = PatternFill("solid", fgColor=Color.BLANCO)
    banner4.alignment = ALIGN_CENTER
    ws2.row_dimensions[2].height = 16

    # ---- KPI Row (row 3-4) ----
    kpi_defs = [
        ("B3", "Total Percepciones", Color.AZUL,
         "B4", '=SUBTOTAL(109,Nomina_Empleados[Sueldo])'),
        ("E3", "Total Deducciones (ISR+IMSS)", Color.ROJO,
         "E4", '=SUBTOTAL(109,Nomina_Empleados[ISR])+SUBTOTAL(109,Nomina_Empleados[IMSS])'),
        ("H3", "ISR del Periodo", Color.AMARILLO,
         "H4", '=SUBTOTAL(109,Nomina_Empleados[ISR])'),
        ("K3", "e.firma Status", Color.VERDE,
         "K4", "VIGENTE"),
    ]

    for title_cell_ref, title, color, val_cell_ref, formula in kpi_defs:
        # Title
        cell_t = ws2[title_cell_ref]
        cell_t.value = title
        cell_t.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell_t.fill = PatternFill("solid", fgColor=color)
        cell_t.alignment = ALIGN_CENTER
        cell_t.border = THIN_BORDER

        # Merge title across 2 cols
        t_row = cell_t.row
        t_col = cell_t.column
        ws2.merge_cells(start_row=t_row, start_column=t_col,
                        end_row=t_row, end_column=t_col + 1)
        ws2.cell(row=t_row, column=t_col + 1).fill = PatternFill("solid", fgColor=color)
        ws2.cell(row=t_row, column=t_col + 1).border = THIN_BORDER

        # Value
        cell_v = ws2[val_cell_ref]
        cell_v.value = formula
        cell_v.font = Font(name="Calibri", bold=True, size=14, color=color)
        cell_v.alignment = ALIGN_CENTER
        cell_v.border = THIN_BORDER
        if "SUBTOTAL" in str(formula):
            cell_v.number_format = FMT_MONEY

        # Merge value across 2 cols
        v_row = cell_v.row
        v_col = cell_v.column
        ws2.merge_cells(start_row=v_row, start_column=v_col,
                        end_row=v_row, end_column=v_col + 1)
        ws2.cell(row=v_row, column=v_col + 1).border = THIN_BORDER

    ws2.row_dimensions[3].height = 25
    ws2.row_dimensions[4].height = 35

    # ---- Filters area (col A-B, rows 6-25) ----
    ws2.merge_cells("A6:B6")
    filt_title = ws2.cell(row=6, column=1, value="FILTROS / SEGMENTADORES")
    filt_title.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    filt_title.fill = PatternFill("solid", fgColor=Color.AZUL)
    filt_title.alignment = ALIGN_CENTER
    ws2.cell(row=6, column=2).fill = PatternFill("solid", fgColor=Color.AZUL)

    filter_labels = ["Periodo:", "Puesto:", "Empleado:"]
    for i, label in enumerate(filter_labels):
        r = 8 + i * 4
        ws2.cell(row=r, column=1, value=label).font = FONT_NORMAL
        ws2.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws2.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=2)
        ws2.cell(row=r + 1, column=1,
                 value="[Insertar segmentador]").font = FONT_SMALL

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12

    # ---- Chart placeholders (rows 6-25, cols C-M) ----
    _make_placeholder(ws2, 6, 15, 3, 8, "Grafico 1: Sueldo por Periodo (linea)")
    _make_placeholder(ws2, 6, 15, 9, 13, "Grafico 2: ISR por Puesto (barras)")
    _make_placeholder(ws2, 17, 25, 3, 13, "Grafico 3: Detalle neto por empleado (tabla o barras)")

    # Remaining column widths
    for c in range(3, 14):
        ws2.column_dimensions[get_column_letter(c)].width = 12

    # Footer
    ws2.merge_cells("A27:M27")
    footer = ws2.cell(row=27, column=1,
                      value="Curso: Excel para Contadores y Administrativos | Israel Castro | 2026")
    footer.font = FONT_SMALL
    footer.alignment = ALIGN_CENTER

    ws2.sheet_properties.tabColor = Color.AZUL

    # ==== Hoja 3: Instrucciones ===========================================
    gen.add_instructions_sheet([
        "SOLUCION DE REFERENCIA -- Este archivo muestra el resultado final del dashboard del Modulo 4.",
        "Abre primero el 09_Layout_Dashboard_Contable.xlsx para construir tu propio dashboard en clase.",
        "La hoja 'Datos_Nomina' tiene 240 registros (20 empleados x 12 meses) como tabla nombrada 'Nomina_Empleados'.",
        "Los datos simulan sueldos mexicanos reales desde salario minimo hasta $55,000/mes.",
        "La hoja 'Dashboard' tiene KPIs con formulas SUBTOTAL que se actualizan con los segmentadores.",
        "PASO 1: Crea Tablas Dinamicas desde Datos_Nomina (una por meses, otra por puestos).",
        "PASO 2: Inserta Segmentadores (Periodo, Puesto, Empleado) vinculados a ambas TDs.",
        "PASO 3: Crea graficos (barras y lineas) y muevalos a la hoja Dashboard.",
        "PASO 4: Protege las hojas y comparte como PDF o Excel protegido.",
    ])

    gen.save()


if __name__ == "__main__":
    build()
