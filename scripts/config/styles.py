"""
Estilos compartidos para openpyxl, reportlab y python-pptx.
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from scripts.config.constants import Color, FONT_TITULO, FONT_CUERPO

# ── openpyxl: Fills ───────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor=Color.AZUL)
FILL_LIGHT = PatternFill("solid", fgColor=Color.FONDO_CLARO)
FILL_MEDIO = PatternFill("solid", fgColor=Color.FONDO_MEDIO)
FILL_VERDE = PatternFill("solid", fgColor=Color.VERDE)
FILL_ROJO = PatternFill("solid", fgColor=Color.ROJO)
FILL_AMARILLO = PatternFill("solid", fgColor=Color.AMARILLO)
FILL_BLANCO = PatternFill("solid", fgColor=Color.BLANCO)

# ── openpyxl: Fonts ──────────────────────────────────────────────
FONT_HEADER = Font(name=FONT_TITULO, bold=True, size=11, color=Color.BLANCO)
FONT_TITULO_XL = Font(name=FONT_TITULO, bold=True, size=14, color=Color.AZUL)
FONT_SUBTITULO = Font(name=FONT_TITULO, bold=True, size=12, color=Color.TEXTO_OSCURO)
FONT_NORMAL = Font(name=FONT_CUERPO, size=11, color=Color.TEXTO_OSCURO)
FONT_SMALL = Font(name=FONT_CUERPO, size=9, color=Color.TEXTO_MEDIO)
FONT_LINK = Font(name=FONT_CUERPO, size=11, color=Color.AZUL, underline="single")

# ── openpyxl: Borders ────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color=Color.GRIS_BORDE),
    right=Side(style="thin", color=Color.GRIS_BORDE),
    top=Side(style="thin", color=Color.GRIS_BORDE),
    bottom=Side(style="thin", color=Color.GRIS_BORDE),
)

# ── openpyxl: Alignment ─────────────────────────────────────────
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# ── openpyxl: Number formats ────────────────────────────────────
FMT_MONEY = '#,##0.00'
FMT_PCT = '0.00%'
FMT_INT = '#,##0'
FMT_DATE = 'DD/MM/YYYY'
FMT_MONEY_4DEC = '#,##0.0000'


def apply_header_style(ws, row, col_start, col_end):
    """Aplica estilo de encabezado a una fila."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def apply_data_style(ws, row, col_start, col_end, money_cols=None, pct_cols=None, date_cols=None):
    """Aplica estilo de datos a una fila."""
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    date_cols = date_cols or []
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_LEFT
        if col in money_cols:
            cell.number_format = FMT_MONEY
            cell.alignment = ALIGN_RIGHT
        elif col in pct_cols:
            cell.number_format = FMT_PCT
            cell.alignment = ALIGN_RIGHT
        elif col in date_cols:
            cell.number_format = FMT_DATE
            cell.alignment = ALIGN_CENTER


def style_title_cell(ws, row, col, text, merge_end_col=None):
    """Crea una celda de título con estilo."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_TITULO_XL
    cell.alignment = ALIGN_LEFT
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)


def style_instruction_cell(ws, row, col, text, merge_end_col=None):
    """Crea una celda de instrucción con estilo."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_SMALL
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)


def auto_width(ws, min_width=10, max_width=40):
    """Ajusta el ancho de columnas basado en contenido."""
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_len = min_width
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        if col_letter:
            ws.column_dimensions[col_letter].width = max_len
